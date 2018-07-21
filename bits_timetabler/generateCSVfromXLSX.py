#! /usr/bin/python
import openpyxl
import sys
import os
import csv
import re

try:
    import progressbar
except ImportError:
    print("Running without loading bar")
    progressbar = range

if len(sys.argv) != 2:
    print("usage {} pathToXlsxFile".format(os.path.basename(__file__)))
    print("see sample.xlsx for exact format that the xlsx file should be in")
    sys.exit(-1)

if not os.path.exists(os.path.basename(sys.argv[1])):
    print("{} does not exist...exiting".format(os.path.basename(sys.argv[1])))
    sys.exit(-1)


print("Generating CSVs from {}".format(sys.argv[1]))
wb = openpyxl.load_workbook(os.path.abspath(sys.argv[1]))
ws = wb.active


print("initial cleaning")

for i in progressbar.iterable(1, ws.max_row+1):
    for k in range(1, 12):
        if ws.cell(i, k).value is None:
            ws.cell(i, k).value = ws.cell(i-1, k).value

wb.save("./GENERATED.xlsx")
print("Initial Cleaning finished (./GENERATED.xlsx)")
print("Generating timetable_data.csv")

last_code = ws.cell(1, 2).value
last_sub_name = ws.cell(1, 3).value


def daysToNums(x):
    # Maps  M to 0
    # 		T to 1
    # 		W to 2
    #  		Th to 3
    # 		F to 4
    # 		S to 5
    my_list = ["M", "T", "W", "Th", "F", "S"]
    return my_list.index(x)


with open("timetable_data.csv", "w") as csvfile:
    fieldnames = ["comCode", "courseCode", "courseTitle", "L", "P", "U",
                  "sectionNo", "instructorName", "instructorId", "department",
                  "roomNo", "days", "hours", "compreDate", "slotType"]
    csvWriter = csv.DictWriter(csvfile, fieldnames)
    csvWriter.writeheader()
    LPUregex = re.compile(r"(\d+)")
    instructorRegex = re.compile(r"^(.*)\((.*)\)$")
    hoursRegex = re.compile(r"(\d)")
    for i in progressbar.iterable(1, ws.max_row+1):
        if last_code != ws.cell(i, 2).value:
            last_code = ws.cell(i, 2).value
            last_sub_name = ws.cell(i, 3).value

        # The following code is hardcoded and extremely brittle
        # it expects the cols of the xlsx sheet to be in a specific order
        # it also expects the cell values in a specific format
        my_dict = {}
        my_dict["comCode"] = ws.cell(i, 1).value
        my_dict["courseCode"] = last_code
        my_dict["courseTitle"] = last_sub_name
        LPU = LPUregex.findall(str(ws.cell(i, 4).value))
        my_dict["L"] = int(LPU[0])
        my_dict["P"] = int(LPU[1])
        my_dict["U"] = int(LPU[2])
        my_dict["sectionNo"] = int(ws.cell(i, 5).value)
        if instructorRegex.match(str(ws.cell(i, 6).value)):
            nameAndId = instructorRegex.search(
                str(ws.cell(i, 6).value)).groups()
        else:
            print("Format :- {name} ({id}) cell row ", i, ' col ', 6)
            print("Exiting..")
            sys.exit(-1)
        my_dict["instructorName"] = nameAndId[0]
        my_dict["instructorId"] = int(re.sub("[^0-9]", "", nameAndId[1]))
        my_dict["department"] = ws.cell(i, 7).value
        my_dict["roomNo"] = ws.cell(i, 8).value
        my_dict["days"] = [*map(daysToNums, str(ws.cell(i, 9).value).split())]
        my_dict["hours"] = [
            int(x)-1 for x in str(ws.cell(i, 10).value).split()]
        my_dict["compreDate"] = ws.cell(i, 11).value
        if ws.cell(i, 3).value != last_sub_name:
            my_dict["slotType"] = ws.cell(i, 3).value
        elif my_dict["L"] > 0:
            my_dict["slotType"] = "Lecture"
        elif my_dict["P"] > 0:
            my_dict["slotType"] = "Practical"
        else:
            my_dict["slotType"] = "Tutorial"
        csvWriter.writerow(my_dict)
